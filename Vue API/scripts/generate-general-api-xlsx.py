#!/usr/bin/env python3
"""Generate general-api.xlsx from general-api.md using the example xlsx as a style template."""

from __future__ import annotations

import html
import re
import shutil
import sys
from copy import copy
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
MD_PATH = ROOT / "general-api.md"
OUTPUT_PATH = ROOT / "general-api.xlsx"
TEMPLATE_GLOB = "4*.xlsx"

STRUCTURE: list[str | tuple[str, str]] = [
    ("section", "Общие утилиты"),
    "version",
    "nextTick()",
    ("section", "Определение компонентов"),
    "defineComponent()",
    "defineAsyncComponent()",
]

CATEGORY_BY_API: dict[str, str] = {
    "version": "Для проверки версии Vue",
    "nextTick()": "Для работы с DOM после обновления",
    "defineComponent()": "Для TypeScript и определения компонентов",
    "defineAsyncComponent()": "Для ленивой загрузки компонентов",
}


def clean_inline(text: str) -> str:
    text = text.strip()
    text = re.sub(r"`([^`]+)`", r"\1", text)
    text = re.sub(r"\*\*([^*]+)\*\*", r"\1", text)
    return html.unescape(text)


def html_to_text(value: str) -> str:
    if not value or value.strip() in {"—", "-"}:
        return ""
    text = value.strip()
    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"</?pre>", "", text, flags=re.IGNORECASE)
    text = re.sub(r"</?code>", "", text, flags=re.IGNORECASE)
    return html.unescape(text).strip()


def parse_links(value: str) -> list[tuple[str, str]]:
    if not value or value.strip() in {"—", "-"}:
        return []
    return re.findall(r"\[([^\]]+)\]\(([^)]+)\)", value)


def split_table_row(line: str) -> list[str]:
    parts = [part.strip() for part in line.strip().split("|")]
    inner = parts[1:-1] if parts and parts[0] == "" else parts
    if len(inner) == 6:
        return inner
    if len(inner) > 6:
        return [
            inner[0],
            inner[1],
            inner[2],
            "|".join(inner[3:-2]),
            inner[-2],
            inner[-1],
        ]
    raise ValueError(f"Unexpected table row with {len(inner)} columns: {line[:120]}")


def parse_md_table(path: Path) -> dict[str, dict[str, object]]:
    rows: dict[str, dict[str, object]] = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line.startswith("|") or line.startswith("|---"):
            continue
        parts = split_table_row(line)
        if parts[0] == "API":
            continue
        api = clean_inline(parts[0])
        rows[api] = {
            "api": api,
            "description": clean_inline(parts[1]),
            "default": clean_inline(parts[2]),
            "example": html_to_text(parts[3]),
            "details": clean_inline(parts[4]),
            "links": parse_links(parts[5]),
        }
    return rows


def format_example(example: str, default: str) -> str:
    default = default.strip()
    if not example:
        if default and default not in {"—", "-"}:
            return f"//по умолчанию {default.strip('`')}"
        return ""
    if default and default not in {"—", "-"} and "по умолчанию" not in example.lower():
        example = f"{example}\n//по умолчанию {default.strip('`')}"
    return example


def format_links(links: list[tuple[str, str]]) -> str:
    if not links:
        return "—"
    return "\n".join(text for text, _url in links)


def apply_style(template_cell, target_cell) -> None:
    if template_cell.has_style:
        target_cell.font = copy(template_cell.font)
        target_cell.fill = copy(template_cell.fill)
        target_cell.border = copy(template_cell.border)
        target_cell.alignment = copy(template_cell.alignment)
        target_cell.number_format = copy(template_cell.number_format)
        target_cell.protection = copy(template_cell.protection)
    target_cell.style = template_cell.style


def estimate_row_height(text: str | None, min_height: float = 15.0) -> float:
    if not text:
        return min_height
    lines = max(1, str(text).count("\n") + 1)
    return max(min_height, min(330.0, lines * 15.0))


def find_template() -> Path:
    matches = sorted(ROOT.glob(TEMPLATE_GLOB))
    if not matches:
        raise FileNotFoundError(f"Template xlsx not found in {ROOT}")
    return matches[0]


def clear_rows_from(ws, start_row: int) -> None:
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)


def write_hyperlink(cell, links: list[tuple[str, str]], template_link_cell) -> None:
    if not links:
        cell.value = "—"
        apply_style(template_link_cell, cell)
        return
    text, url = links[0]
    cell.value = text
    cell.hyperlink = url
    apply_style(template_link_cell, cell)
    if len(links) > 1:
        extra = "\n".join(f"{label} ({href})" for label, href in links[1:])
        cell.value = f"{text}\n{extra}"


def main() -> int:
    if not MD_PATH.exists():
        print(f"Missing markdown source: {MD_PATH}", file=sys.stderr)
        return 1

    md_rows = parse_md_table(MD_PATH)
    missing = [item for item in STRUCTURE if isinstance(item, str) and item not in md_rows]
    if missing:
        print("Missing APIs in markdown:", ", ".join(missing), file=sys.stderr)
        return 1

    template_path = find_template()
    shutil.copy2(template_path, OUTPUT_PATH)
    wb = load_workbook(OUTPUT_PATH)
    ws = wb.active
    ws.title = "General API"

    # Remove Application-specific index rows from template (keep header row 14).
    for row in range(2, 14):
        for col in range(1, 7):
            ws.cell(row, col).value = None

    style_section = {col: ws.cell(15, col) for col in range(1, 7)}
    style_data = {col: ws.cell(16, col) for col in range(1, 7)}

    clear_rows_from(ws, 15)
    current_row = 15

    for item in STRUCTURE:
        if isinstance(item, tuple):
            _, title = item
            for col in range(1, 7):
                cell = ws.cell(current_row, col)
                apply_style(style_section[col], cell)
                cell.value = title if col == 1 else None
            ws.row_dimensions[current_row].height = 15
            current_row += 1
            continue

        api = item
        row_data = md_rows[api]
        example = format_example(str(row_data["example"]), str(row_data["default"]))
        links = row_data["links"]
        assert isinstance(links, list)

        values = [
            api,
            row_data["description"],
            CATEGORY_BY_API.get(api, "—"),
            example or "—",
            row_data["details"] or "—",
            format_links(links),
        ]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(current_row, col)
            apply_style(style_data[col], cell)
            if col == 6:
                write_hyperlink(cell, links, style_data[6])
            else:
                cell.value = value

        ws.row_dimensions[current_row].height = estimate_row_height(
            max((str(v) for v in values if v), key=len, default="")
        )
        current_row += 1

    ws.sheet_view.topLeftCell = "A1"
    ws.sheet_view.selection[0].sqref = "A15"

    wb.save(OUTPUT_PATH)
    data_count = sum(1 for item in STRUCTURE if isinstance(item, str))
    print(f"Generated {OUTPUT_PATH.name} ({data_count} APIs, template: {template_path.name})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
